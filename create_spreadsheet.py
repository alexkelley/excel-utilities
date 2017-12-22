#! /usr/bin/env python
# -*- coding: utf-8 -*-

import time
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from operator import itemgetter
import datetime
import pprint

            
def create_spreadsheet(data_list, worksheet_title, filename):
    '''
    Parameters:
    - data_list is a list of row data each in its own list (works with tuples too)
    - worksheet_title is a string
    - filename is a string.  Must include absolute path and file extension

    Returns:
    - True if sucessful; False if not
    '''
    columnHeader = NamedStyle(name='columnHeader')
    columnHeader.font = Font(name='Calibri', bold=True)
    columnHeader.alignment = Alignment(horizontal="center")
    columnHeader.fill = PatternFill(fill_type='solid',
                                    start_color='99ccff', end_color='99ccff')
    
    dollarFormat = NamedStyle(name='dollarFormat')
    dollarFormat.number_format = '$#,##0.00'
    dollarFormat.alignment = Alignment(horizontal="right")
    dollarFormat.font = Font(name='Calibri')

    try:
        wb = openpyxl.Workbook()

        ## use for automatically creating additional worksheets
        # for i in range(len(data.keys())-1):
        #     wb.create_sheet()

        ws = wb.worksheets[0]
        ws.title = worksheet_title
        ws.column_dimensions['A'].width = 17

        row = 1
        for data_row in data_list:
            for i, data_point in enumerate(data_row):
                ws.cell(row=row, column=i+1).value = data_point
                #ws.cell(row=row, column=i+1).style = dollarFormat
                #ws.cell(row=row, column=i+1).alignment = Alignment(horizontal="center")

            row += 1

        wb.save(filename)
        flag = True

    except:
        flag = False
    
    return flag

##################
# Function Calls #
##################

def main():
    start_time = time.time()

    timestamp = datetime.datetime.strftime(
        datetime.datetime.now(), '%Y-%m-%dT%H-%M-%S')
    path = '/home/'
    file_format = 'xslx'
    filename = '{0}working_filename_{1}.{2}'.format(path, timestamp, file_format)

    data_list = []
    worksheet_title = 'Working Title'

    create_spreadsheet(data_list, worksheet_title, filename)

    end_time = time.time()
    print('\nFile saved:\n{}'.format(filename))
    print('Elapsed time: {:.2f} seconds\n'.format(end_time - start_time))

if __name__ == '__main__':
    main()
